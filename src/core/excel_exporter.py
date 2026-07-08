"""Excel export: openpyxl → xlsx artifact. See [[Artifacts]] (spec §5.4).

Formatting: bold header, auto column width, money number format, frozen header
row. Large exports (> ROW_WARN_THRESHOLD) should be flagged in the text answer
by the caller (the exporter still writes the file).
"""

from __future__ import annotations

import logging
import uuid
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import get_settings

logger = logging.getLogger(__name__)

# Excel custom number format for money: thousands-grouped integer.
MONEY_FORMAT = "#,##0"
ROW_WARN_THRESHOLD = 100_000
MAX_COL_WIDTH = 60
HEADER_FILL = "DDDDDD"


def _is_numeric_column(rows: list[dict], col: str) -> bool:
    """True if the column's first non-null value is a (non-bool) number."""
    for r in rows:
        v = r.get(col)
        if v is None:
            continue
        return isinstance(v, (int, float)) and not isinstance(v, bool)
    return False


def export_to_excel(
    rows: list[dict],
    columns: list[str] | None = None,
    *,
    out_dir: Path | None = None,
    sheet_name: str = "Данные",
) -> Path:
    """Write ``rows`` to a formatted .xlsx and return its path."""
    if columns is None:
        columns = list(rows[0].keys()) if rows else []
    out_dir = out_dir or get_settings().artifacts_path

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet-name limit

    # Header row.
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(vertical="center")

    # Data rows.
    for r in rows:
        ws.append([r.get(c) for c in columns])

    numeric_cols = {c for c in columns if _is_numeric_column(rows, c)}

    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        # Money number format for numeric columns (skip the header row).
        if col in numeric_cols:
            for cell in ws[letter][1:]:
                cell.number_format = MONEY_FORMAT
        # Auto column width from the widest cell in the column.
        widest = max([len(str(col))] + [len(str(r.get(col, ""))) for r in rows] or [0])
        ws.column_dimensions[letter].width = min(widest + 2, MAX_COL_WIDTH)

    ws.freeze_panes = "A2"  # freeze the header row

    path = out_dir / f"export_{uuid.uuid4().hex[:12]}.xlsx"
    wb.save(path)
    logger.info("Excel saved: %s (%d rows, %d cols)", path, len(rows), len(columns))
    return path
