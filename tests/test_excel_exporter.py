"""Unit tests for the Excel exporter."""

from __future__ import annotations

from openpyxl import load_workbook

from src.core.excel_exporter import MONEY_FORMAT, export_to_excel

ROWS = [
    {"city": "Москва", "revenue": 1500000, "share": 0.42},
    {"city": "Казань", "revenue": 750000, "share": 0.21},
]


def test_export_creates_valid_xlsx(tmp_path) -> None:
    path = export_to_excel(ROWS, out_dir=tmp_path)
    assert path.exists() and path.suffix == ".xlsx"

    wb = load_workbook(path)
    ws = wb.active
    # Header present and bold.
    assert [c.value for c in ws[1]] == ["city", "revenue", "share"]
    assert ws["A1"].font.bold is True
    # Header row frozen.
    assert ws.freeze_panes == "A2"
    # Values round-trip.
    assert ws["A2"].value == "Москва"
    assert ws["B2"].value == 1500000


def test_numeric_columns_get_money_format(tmp_path) -> None:
    path = export_to_excel(ROWS, out_dir=tmp_path)
    ws = load_workbook(path).active
    # revenue (numeric) formatted; city (string) not.
    assert ws["B2"].number_format == MONEY_FORMAT
    assert ws["A2"].number_format != MONEY_FORMAT


def test_explicit_columns_subset_and_order(tmp_path) -> None:
    path = export_to_excel(ROWS, columns=["revenue", "city"], out_dir=tmp_path)
    ws = load_workbook(path).active
    assert [c.value for c in ws[1]] == ["revenue", "city"]
    assert ws["A2"].value == 1500000


def test_empty_rows_writes_header_only(tmp_path) -> None:
    path = export_to_excel([], columns=["a", "b"], out_dir=tmp_path)
    ws = load_workbook(path).active
    assert [c.value for c in ws[1]] == ["a", "b"]
    assert ws.max_row == 1
